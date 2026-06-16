"""
Genera PAUSIA_mapa_frecuencia_atomico_completo_final.xlsx a partir de PDFs reales.

Requisitos previstos:
  python scripts/generate_pausia_frequency_excel.py

Dependencias recomendadas:
  pandas openpyxl pdfplumber pypdf

Nota de entorno: este repositorio puede no tener Python instalado en la maquina
local. En ese caso, use el runner equivalente:
  node scripts/generate_pausia_frequency_excel.mjs
"""

from __future__ import annotations

import datetime as dt
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "PAUSIA_mapa_frecuencia_atomico_completo_final.xlsx"
COMMUNITIES = ["Madrid", "Cataluña"]


def normalize(text: str) -> str:
    text = unicodedata.normalize("NFD", text or "")
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text.lower()).strip()


def level(probability: float) -> str:
    if probability >= 70:
        return "Crítica / Presencia Continua"
    if probability >= 50:
        return "Muy Alta"
    if probability >= 30:
        return "Alta"
    if probability >= 15:
        return "Media"
    if probability >= 5:
        return "Baja"
    if probability > 0:
        return "Marginal / Testimonial"
    return "No Detectada en histórico"


def taxonomy() -> list[dict]:
    items = [
        ("Matemáticas II", "Álgebra", "Matrices y determinantes", ["matriz", "matrices", "determinante", "rango"]),
        ("Matemáticas II", "Álgebra", "Sistemas lineales", ["sistema de ecuaciones", "discuta el sistema", "compatible"]),
        ("Matemáticas II", "Análisis", "Límites y continuidad", ["limite", "continuidad", "asintota"]),
        ("Matemáticas II", "Análisis", "Derivadas y optimización", ["derivada", "maximo", "minimo", "optimizar"]),
        ("Matemáticas II", "Análisis", "Integrales y áreas", ["integral", "area encerrada", "primitiva"]),
        ("Matemáticas II", "Geometría", "Rectas y planos", ["recta", "plano", "posicion relativa"]),
        ("Matemáticas II", "Probabilidad", "Distribución binomial y normal", ["binomial", "normal", "probabilidad"]),
        ("Historia de España", "Antiguo Régimen", "Sociedad estamental y reformismo borbónico", ["antiguo regimen", "borbones", "ilustracion", "reformismo borbonico"]),
        ("Historia de España", "Crisis liberal", "Guerra de Independencia y Constitución de 1812", ["guerra de independencia", "constitucion de 1812", "cadiz"]),
        ("Historia de España", "Siglo XIX", "Fernando VII e Isabel II", ["fernando vii", "isabel ii", "regencias"]),
        ("Historia de España", "Siglo XIX", "Sexenio Democrático y Restauración", ["sexenio democratico", "restauracion", "canovas"]),
        ("Historia de España", "Siglo XX", "Segunda República", ["segunda republica", "bienio", "frente popular"]),
        ("Historia de España", "Siglo XX", "Guerra Civil", ["guerra civil", "sublevacion", "bando republicano"]),
        ("Historia de España", "Siglo XX", "Franquismo", ["franquismo", "franco", "autarquia", "desarrollismo"]),
        ("Historia de España", "Democracia", "Transición y Constitución de 1978", ["transicion", "constitucion de 1978", "adolfo suarez"]),
        ("Historia de la Filosofía", "Antigua", "Platón: teoría de las Ideas", ["platon", "ideas", "mito de la caverna"]),
        ("Historia de la Filosofía", "Antigua", "Aristóteles: sustancia y causas", ["aristoteles", "sustancia", "causas"]),
        ("Historia de la Filosofía", "Moderna", "Descartes: método y cogito", ["descartes", "cogito", "duda metodica"]),
        ("Historia de la Filosofía", "Moderna", "Kant: conocimiento y ética", ["kant", "imperativo categorico", "critica de la razon"]),
        ("Historia de la Filosofía", "Contemporánea", "Nietzsche: crítica de la moral", ["nietzsche", "moral", "superhombre"]),
        ("Historia de la Filosofía", "Contemporánea", "Ortega: razón vital", ["ortega", "razon vital", "yo soy yo"]),
        ("Lengua Castellana y Literatura", "Comentario", "Tema, resumen y comentario crítico", ["tema del texto", "resumen", "comentario de texto"]),
        ("Lengua Castellana y Literatura", "Lengua", "Sintaxis de oración compleja", ["analisis sintactico", "oracion compuesta", "subordinada"]),
        ("Lengua Castellana y Literatura", "Lengua", "Morfología y formación de palabras", ["morfologia", "derivacion", "parasintesis"]),
        ("Lengua Castellana y Literatura", "Literatura", "Generación del 98 y Modernismo", ["generacion del 98", "modernismo", "unamuno"]),
        ("Lengua Castellana y Literatura", "Literatura", "Novecentismo, vanguardias y 27", ["generacion del 27", "vanguardias", "novecentismo"]),
        ("Inglés", "Reading", "Comprensión global y específica", ["reading", "according to the text", "true or false"]),
        ("Inglés", "Use of English", "Word formation", ["word formation", "complete the sentences", "rewrite"]),
        ("Inglés", "Use of English", "Connectors and grammar transformations", ["connectors", "reported speech", "passive voice"]),
        ("Inglés", "Writing", "Opinion essay / email", ["write", "essay", "email", "opinion"]),
        ("Biología", "Bioquímica", "Biomoléculas y metabolismo", ["glucidos", "lipidos", "proteinas", "metabolismo"]),
        ("Biología", "Genética", "Herencia mendeliana y genética molecular", ["adn", "arn", "mendel", "mutacion"]),
        ("Biología", "Inmunología", "Sistema inmunitario", ["anticuerpo", "linfocito", "inmunidad"]),
        ("Biología", "Ecología", "Ecosistemas y evolución", ["ecosistema", "evolucion", "seleccion natural"]),
        ("Física", "Campo gravitatorio", "Gravitación y órbitas", ["campo gravitatorio", "orbita", "ley de gravitacion"]),
        ("Física", "Electromagnetismo", "Campo eléctrico y magnético", ["campo electrico", "campo magnetico", "induccion"]),
        ("Física", "Ondas", "Movimiento ondulatorio", ["onda", "interferencia", "difraccion"]),
        ("Física", "Física moderna", "Cuántica y nuclear", ["efecto fotoelectrico", "nuclear", "radiactividad"]),
        ("Química", "Equilibrio", "Constantes Kc/Kp y Le Châtelier", ["kc", "kp", "le chatelier", "equilibrio quimico"]),
        ("Química", "Ácido-base", "pH, Ka/Kb e hidrólisis", ["ph", "ka", "kb", "hidrolisis", "acido base"]),
        ("Química", "Redox", "Ajuste redox y potenciales", ["redox", "potencial", "electrolisis", "oxidacion"]),
        ("Química", "Orgánica", "Isomería y reacciones orgánicas", ["isomeria", "organica", "alcohol", "ester"]),
        ("Química", "Termoquímica", "Entalpía y energía libre", ["entalpia", "energia libre", "gibbs"]),
        ("Química", "Cinética", "Velocidad de reacción", ["cinetica", "velocidad de reaccion", "energia de activacion"]),
    ]
    rows = []
    idx = 1
    for community in COMMUNITIES:
        for matter, block, concept, synonyms in items:
            rows.append({
                "concept_id": f"MC-{idx:04d}",
                "comunidad": community,
                "materia": matter,
                "bloque": block,
                "microconcepto": concept,
                "sinonimos": synonyms,
                "peso_producto": 90 if matter in {"Matemáticas II", "Historia de España", "Inglés"} else 75,
                "ajuste_temario": 90,
            })
            idx += 1
    return rows


def extract_pdf_text(path: Path) -> tuple[str, str]:
    errors = []
    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(path) as pdf:
            return "\n".join((page.extract_text() or "") for page in pdf.pages), "pdfplumber"
    except Exception as exc:  # pragma: no cover - audit fallback
        errors.append(f"pdfplumber:{exc}")
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        return "\n".join((page.extract_text() or "") for page in reader.pages), "pypdf"
    except Exception as exc:
        errors.append(f"pypdf:{exc}")
    raise RuntimeError("; ".join(errors) or "No hay extractor PDF disponible")


def detect_exam(path: Path) -> dict:
    name = normalize(path.as_posix())
    year = int(re.search(r"(20\d{2}|19\d{2})", name).group(1)) if re.search(r"(20\d{2}|19\d{2})", name) else 0
    subject = "Química" if "quimica" in name else "Historia de España" if "historia" in name else "Desconocida"
    call = "Extraordinaria" if "extraordinaria" in name else "Ordinaria"
    variant = "lunes" if "lunes" in name else "martes" if "martes" in name else ""
    community = "Cataluña" if "cataluna" in name or "catalunya" in name else "Madrid"
    return {"materia": subject, "comunidad": community, "año": year, "convocatoria": call, "variante": variant}


def split_segments(text: str) -> list[str]:
    parts = re.split(r"(?=\b(?:Pregunta|Ejercicio|Exercici|Option|Opcion|Apartado)\s+\d|\n\s*\d+\.)", text)
    parts = [re.sub(r"\s+", " ", p).strip() for p in parts if len(p.strip()) > 80]
    return parts[:40] or [re.sub(r"\s+", " ", text).strip()[:1200]]


def main() -> None:
    import pandas as pd  # type: ignore
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill  # type: ignore

    concepts = taxonomy()
    pdfs = sorted((ROOT / "public").rglob("*.pdf"))
    audit, exams, base_rows = [], [], []
    presence: dict[tuple[str, str, str], set[str]] = defaultdict(set)

    for path in pdfs:
        meta = detect_exam(path)
        exam_id = f"{meta['materia']}|{meta['comunidad']}|{meta['año']}|{meta['convocatoria']}|{path.name}"
        try:
            text, extractor = extract_pdf_text(path)
            ntext = normalize(text)
            matched = []
            for concept in concepts:
                if concept["materia"] != meta["materia"] or concept["comunidad"] != meta["comunidad"]:
                    continue
                if any(normalize(s) in ntext for s in concept["sinonimos"]):
                    matched.append(concept)
                    presence[(concept["comunidad"], concept["materia"], concept["concept_id"])].add(exam_id)
            exams.append({**meta, "exam_id": exam_id, "pdf": str(path.relative_to(ROOT)), "chars": len(text), "extractor": extractor})
            segments = split_segments(text)
            if matched:
                for concept in matched:
                    base_rows.append({**meta, "ID Examen": exam_id, "Materia": meta["materia"], "Comunidad": meta["comunidad"], "Año": meta["año"], "Convocatoria": meta["convocatoria"], "Variante": meta["variante"], "Microconcepto Detectado": concept["microconcepto"], "Bloque": concept["bloque"], "Pregunta Tipo Textual": segments[0][:900], "PDF de Origen": str(path.relative_to(ROOT)), "Respuesta Esperada": "No disponible en PDF; usar solucion orientativa/criterios si existen en datos TS."})
            else:
                base_rows.append({**meta, "ID Examen": exam_id, "Materia": meta["materia"], "Comunidad": meta["comunidad"], "Año": meta["año"], "Convocatoria": meta["convocatoria"], "Variante": meta["variante"], "Microconcepto Detectado": "Sin coincidencia controlada", "Bloque": "", "Pregunta Tipo Textual": segments[0][:900], "PDF de Origen": str(path.relative_to(ROOT)), "Respuesta Esperada": "No disponible"})
            audit.append({"tipo": "PDF_PROCESADO", "ruta": str(path.relative_to(ROOT)), "detalle": f"{extractor}; chars={len(text)}; conceptos={len(matched)}"})
        except Exception as exc:
            audit.append({"tipo": "ERROR_PDF", "ruta": str(path.relative_to(ROOT)), "detalle": str(exc)})

    denominators = Counter((exam["comunidad"], exam["materia"]) for exam in exams)
    ranking = []
    for concept in concepts:
        denom = denominators[(concept["comunidad"], concept["materia"])]
        hits = len(presence[(concept["comunidad"], concept["materia"], concept["concept_id"])])
        prob = round((hits / denom * 100) if denom else 0, 2)
        score = round(prob * 0.60 + concept["peso_producto"] * 0.25 + concept["ajuste_temario"] * 0.15, 2)
        ranking.append({
            "Concept ID": concept["concept_id"], "Comunidad": concept["comunidad"], "Materia": concept["materia"],
            "Bloque": concept["bloque"], "Microconcepto atómico": concept["microconcepto"],
            "Apariciones": hits, "N Exámenes Base Materia": denom, "Probabilidad Empírica %": prob,
            "Nivel PAUSIA": level(prob), "Peso Estratégico Producto": concept["peso_producto"],
            "Ajuste Temario": concept["ajuste_temario"], "Score PAUSIA": score,
            "Sinónimos / patrones": "; ".join(concept["sinonimos"]),
        })

    ranking_df = pd.DataFrame(ranking).sort_values(["Score PAUSIA", "Probabilidad Empírica %"], ascending=False)
    base_df = pd.DataFrame(base_rows)
    audit_df = pd.DataFrame(audit + [{"tipo": "ENTORNO", "ruta": "", "detalle": "Si no hay documento oficial de temario adjunto, se usa temario controlado del prompt y docs/camino-pau."}])
    temario_df = ranking_df[["Concept ID", "Comunidad", "Materia", "Bloque", "Microconcepto atómico", "Nivel PAUSIA"]].copy()

    panel_df = pd.DataFrame([
        ["Generado", dt.datetime.now().isoformat(timespec="seconds")],
        ["PDFs encontrados", len(pdfs)],
        ["PDFs procesados", len(exams)],
        ["Microconceptos en temario", len(concepts)],
        ["Regla probabilidad", "exámenes con aparición / total exámenes materia-comunidad * 100"],
        ["Regla score", "60% probabilidad + 25% peso estratégico + 15% ajuste temario"],
    ], columns=["Métrica", "Valor"])
    teoria_df = temario_df.assign(Trigger="Necesita teoría desplegable si probabilidad >= 15% o score >= 65", Secuencia="Definición -> ejemplo PAU -> plantilla -> práctica")
    plantillas_df = pd.DataFrame([
        ["Comentario histórico", "Identificar fuente, contexto, idea principal, desarrollo y conclusión."],
        ["Problema científico", "Datos, fórmula, sustitución, unidades, interpretación."],
        ["Writing Inglés", "Plan, tesis, conectores, desarrollo, cierre y revisión gramatical."],
        ["Filosofía", "Tesis del autor, conceptos clave, relación con sistema y valoración crítica."],
    ], columns=["Plantilla", "Criterio operativo"])
    cambios_df = pd.DataFrame(exams)[["materia", "comunidad", "año", "convocatoria", "variante", "pdf"]].drop_duplicates() if exams else pd.DataFrame()
    if not cambios_df.empty:
        cambios_df["Cambio formato detectado"] = cambios_df["variante"].apply(lambda v: "Variante/día explícito" if v else "Revisión manual")
    prioridad_df = ranking_df.head(40)[["Comunidad", "Materia", "Bloque", "Microconcepto atómico", "Score PAUSIA", "Nivel PAUSIA"]].copy()
    prioridad_df["Prioridad MVP"] = range(1, len(prioridad_df) + 1)

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        panel_df.to_excel(writer, "00_PANEL", index=False)
        ranking_df.to_excel(writer, "01_RANKING_PAUSIA", index=False)
        base_df.to_excel(writer, "02_BASE_EXAMENES", index=False)
        temario_df.to_excel(writer, "03_TEMARIO_MAPA", index=False)
        teoria_df.to_excel(writer, "04_TEORIA_DESPLEGABLE", index=False)
        plantillas_df.to_excel(writer, "05_PLANTILLAS", index=False)
        cambios_df.to_excel(writer, "06_CAMBIOS_FORMATO", index=False)
        prioridad_df.to_excel(writer, "07_PRIORIDAD_MVP", index=False)
        audit_df.to_excel(writer, "_AUDIT", index=False)

    wb = load_workbook(OUTPUT)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        for col in ws.columns:
            width = min(55, max(12, max(len(str(c.value or "")) for c in col[:60]) + 2))
            ws.column_dimensions[col[0].column_letter].width = width
    wb.save(OUTPUT)
    print(f"OK {OUTPUT}")
    print(f"PDFs encontrados={len(pdfs)} procesados={len(exams)} conceptos={len(concepts)} filas_base={len(base_df)}")


if __name__ == "__main__":
    main()
